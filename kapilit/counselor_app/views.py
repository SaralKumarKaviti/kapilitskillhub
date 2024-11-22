from django.shortcuts import render,redirect,HttpResponse
from django.http import JsonResponse
from counselor_app.models import Manager, Counselor, Role
from django.db import connection,DatabaseError, IntegrityError
from django.conf import settings
from django.contrib import messages
import datetime
import razorpay
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from datetime import datetime, timedelta
from django.core.paginator import Paginator
import openpyxl
from openpyxl.styles import Font


def generate_enrollment_id(course_name, date):
    courses_data = {
        'Python Full Stack': 'KIHPY',
        'Java Full Stack': 'KIHJA',
        'Data Science': 'KIHDS',
        'Digital Marketing': 'KIHDM',
        'UI/UX': 'KIHUI'
    }

    data = courses_data.get(course_name, 'UNKNOWN')
    date_str = date.strftime("%d%m%Y")

    with connection.cursor() as cur:
        # Use DATE() function in MySQL to get the date part
        sequence_query = "SELECT COUNT(*) FROM counselor_app_studentenrollment WHERE course_name = %s AND DATE(enrolled_on) = %s"
        cur.execute(sequence_query, [course_name, date.date()])
        count = cur.fetchone()[0]

    sequential_number = count + 1
    enrollment_id = f"{data}{date_str}{sequential_number:04d}"
    return enrollment_id

def employee_login(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        employee_id = request.POST.get('employee_id').upper()        
        try:
            with connection.cursor() as cur:
                query = "SELECT id FROM counselor_app_role WHERE email = %s AND employee_id = %s"
                cur.execute(query, [email, employee_id])
                employee_data = cur.fetchone()
                
            if employee_data:
                request.session['emp_role_id'] = employee_data[0]
                return redirect('employee_dashboard_page')
            else:
                error = "Invalid Email or Employee ID!"
        except Exception as e:
            error = f"An error occurred during login: {str(e)}"

        return render(request, 'employee/employee_login.html', {'error': error})
        
    return render(request, 'employee/employee_login.html')


def employee_dashboard_page(request):
    try:
        emp_id = request.session.get('emp_role_id')
        if not emp_id:
            return render(request, 'employee/employee_login.html')

        try:
            with connection.cursor() as cur:
                cur.execute("SELECT name, email FROM counselor_app_role WHERE id = %s", [emp_id])
                employee_data = cur.fetchone()
                
            if not employee_data:
                messages.error(request, "Employee details not found.")
                return render(request, 'employee/employee_dashboard.html')
                
            employee_details = {
                "employee_name": employee_data[0],
                "employee_email": employee_data[1],
            }
        except Exception as e:
            messages.error(request, f"Error fetching employee details: {str(e)}")
            return render(request, 'employee/employee_dashboard.html')

        # Get filtering parameters from POST
        course_name_filter = request.POST.get('course_name')
        date_filter = request.POST.get('date_filter')
        name_filter = request.POST.get('name_filter')
        email_filter = request.POST.get('email_filter')
        mobile_filter = request.POST.get('mobile_filter')
        enrollment_id_filter = request.POST.get('enrollment_id_filter')
        
        start_date, end_date = None, None
        params = [emp_id]

        # Prepare the base student query
        student_query = """
            SELECT se.id, se.first_name, se.last_name, se.email, se.mobile, se.location,
                   se.mode_of_attending, se.qualification, se.branch, se.course_name,
                   se.course_amount, se.discount_amount, se.total_amount, se.gender,
                   se.education_status, se.passed_year, se.marks, se.current_year,
                   se.enrolled_on, se.enrollment_id, se.id AS enrolled_id, se.registration_fee, 
                   rp.payment_mode, rp.register_payment_status, rp.paymented_on, 
                   rp.payment_by, rp.transaction_id
            FROM counselor_app_studentenrollment se
            LEFT JOIN counselor_app_registrationpaymentdetails rp
            ON se.id = rp.student_id
            WHERE se.counselor_id = %s
        """

        if course_name_filter:
            student_query += " AND se.course_name = %s"
            params.append(course_name_filter)

        try:
            today = datetime.now().date()
            if date_filter:
                if date_filter == 'today':
                    start_date, end_date = today, today + timedelta(days=1)
                elif date_filter == 'yesterday':
                    start_date, end_date = today - timedelta(days=1), today
                elif date_filter == 'last_3_days':
                    start_date = today - timedelta(days=3)
                elif date_filter == 'last_7_days':
                    start_date = today - timedelta(days=7)
                elif date_filter == 'last_month':
                    start_date = today - timedelta(days=30)
                elif date_filter == 'custom_date':
                    start_date = request.POST.get('start_date')
                    end_date = request.POST.get('end_date')
                    if start_date and end_date:
                        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
                        end_date = datetime.strptime(end_date, "%Y-%m-%d").date() + timedelta(days=1)
                    else:
                        messages.error(request, "Please provide both start and end dates for custom date filter.")
                        return render(request, 'employee/employee_dashboard.html', {'employee_details': employee_details})

                if start_date:
                    student_query += " AND se.enrolled_on >= %s"
                    params.append(start_date)
                if end_date:
                    student_query += " AND se.enrolled_on < %s"
                    params.append(end_date)
        except Exception as e:
            messages.error(request, f"Error applying date filter: {str(e)}")
            return render(request, 'employee/employee_dashboard.html', {'employee_details': employee_details})

        if name_filter:
            student_query += " AND (se.first_name LIKE %s OR se.last_name LIKE %s)"
            params.extend([f"%{name_filter}%", f"%{name_filter}%"])
        if email_filter:
            student_query += " AND se.email LIKE %s"
            params.append(f"%{email_filter}%")
        if mobile_filter:
            student_query += " AND se.mobile LIKE %s"
            params.append(f"%{mobile_filter}%")
        if enrollment_id_filter:
            student_query += " AND se.enrollment_id LIKE %s"
            params.append(f"%{enrollment_id_filter}%")

        try:
            with connection.cursor() as cur:
                cur.execute(student_query, params)
                enrolled_students = cur.fetchall()
        except Exception as e:
            messages.error(request, f"Error fetching student enrollment data: {str(e)}")
            return render(request, 'employee/employee_dashboard.html', {'employee_details': employee_details})

        student_enrollment_details = []
        success_payment_count = 0
        for index, student in enumerate(enrolled_students, start=1):
            register_payment_status = student[23]
            if register_payment_status == "success":
                success_payment_count += 1
            student_enrollment_details.append({
                "s_no": index,
                "id": student[0],
                "first_name": student[1],
                "last_name": student[2],
                "email": student[3],
                "mobile": student[4],
                "location": student[5],
                "mode_of_attending": student[6],
                "qualification": student[7],
                "branch": student[8],
                "course_name": student[9],
                "course_amount": student[10],
                "discount_amount": student[11],
                "total_amount": student[12],
                "gender": student[13],
                "education_status": student[14],
                "passed_year": student[15],
                "marks": student[16],
                "current_year": student[17],
                "enrolled_on": student[18],
                "enrollment_id": student[19],
                "registration_fee": student[20],
                "payment_mode": student[21],
                "register_payment_status": register_payment_status,
                "paymented_on": student[23],
                "payment_by": student[24],
                "transaction_id": student[25]
            })

        if request.GET.get('download') == 'excel':
            return download_enrollment_data(student_enrollment_details, course_name_filter, date_filter)

        paginator = Paginator(student_enrollment_details, 3)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        enrolled_student_count = len(student_enrollment_details)

        final_data = {
            'employee_details': employee_details,
            'student_enrollment_details': page_obj,
            'course_name_filter': course_name_filter,
            'date_filter': date_filter,
            'name_filter': name_filter,
            'email_filter': email_filter,
            'mobile_filter': mobile_filter,
            'enrollment_id_filter': enrollment_id_filter,
            'enrolled_student_count': enrolled_student_count,
            'success_payment_count': success_payment_count
        }

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {str(e)}")
        final_data = {'employee_details': employee_details} if 'employee_details' in locals() else {}

    return render(request, 'employee/employee_dashboard.html', final_data)


def download_enrollment_data(student_enrollment_details, course_name_filter, date_filter):
    try:
        headers = [
            "First Name", "Last Name", "Email", "Mobile", "Location", "Mode of Attending",
            "Qualification", "Branch", "Course Name", "Course Amount", "Discount Amount",
            "Total Amount", "Gender", "Education Status", "Passed Year", "Marks",
            "Current Year", "Enrolled On", "Enrollment ID", "Registration Fee",
            "Payment Mode", "Register Payment Status", "Paymented On",
            "Payment By", "Transaction ID"
        ]
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Enrollment Data'

        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header).font = Font(bold=True)
        for row_num, student in enumerate(student_enrollment_details, 2):
            data = [
                student["first_name"],
                student["last_name"],
                student["email"],
                student["mobile"],
                student["location"],
                student["mode_of_attending"],
                student["qualification"],
                student["branch"],
                student["course_name"],
                student["course_amount"],
                student["discount_amount"],
                student["total_amount"],
                student["gender"],
                student["education_status"],
                student["passed_year"],
                student["marks"],
                student["current_year"],
                student["enrolled_on"],
                student["enrollment_id"],
                student["registration_fee"],
                student["payment_mode"],
                student["register_payment_status"],
                student["paymented_on"],
                student["payment_by"],
                student["transaction_id"]
            ]
            
            for col_num, cell_value in enumerate(data, 1):
                sheet.cell(row=row_num, column=col_num, value=cell_value)

        filename = 'Filtered_Enrollment_Data.xlsx' if course_name_filter or date_filter else 'Unfiltered_Enrollment_Data.xlsx'
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        workbook.save(response)
        return response

    except Exception as e:
        print(f"Error occurred while generating the Excel file: {e}")
        messages.error(request, f"An error occurred while generating the enrollment data: {e}")
        return render(request, 'employee/employee_dashboard.html')

def employee_logout(request):
    try:
        emp_id = request.session.get('emp_role_id')
        if emp_id:
            request.session.pop(emp_id, None)
            messages.success(request, "Successfully logged out!")
        else:
            messages.warning(request, "No active session found.")
        
        return redirect('employee_login')

    except Exception as e:
        print(f"Error occurred during logout: {e}")
        messages.error(request, f"An error occurred while logging out: {e}")
        return redirect('employee_dashboard')

def manager_login(request):
    try:
        if request.method == 'POST':
            email = request.POST.get('email')
            employee_id = request.POST.get('employee_id').upper()

            with connection.cursor() as cur:
                query = "SELECT id FROM counselor_app_manager WHERE email = %s AND employee_id = %s"
                cur.execute(query, [email, employee_id])
                manager_data = cur.fetchone()
                #print(manager_data)

            if manager_data:
                request.session['manager_id'] = manager_data[0]
                return redirect('manager_dashboard_page')
            else:
                error = "Invalid Email or Employee ID!"
                return render(request, 'manager/manager_login.html', {'error': error})

        return render(request, 'manager/manager_login.html')

    except Exception as e:
        print(f"Error occurred during manager login: {e}")
        messages.error(request, f"An error occurred during login: {e}")
        return redirect('manager_login')



def manager_dashboard_page(request):
    try:
        manager_id = request.session.get('manager_id')
        if not manager_id:
            return render(request, 'manager/manager_login.html')

        # Fetch manager details
        with connection.cursor() as cur:
            manager_query = "SELECT name, email, employee_id FROM counselor_app_manager WHERE id = %s"
            cur.execute(manager_query, [manager_id])
            manager_data = cur.fetchone()

        if manager_data:
            manager_details = {
                "manager_name": manager_data[0],
                "manager_email": manager_data[1],
                "manager_empid": manager_data[2]
            }

        # Get filters from POST request
        course_name_filter = request.POST.get('course_name')
        date_filter = request.POST.get('date_filter')
        enrollment_id_filter = request.POST.get('enrollment_id')
        counselor_name_filter = request.POST.get('counselor_name')
        mobile_filter = request.POST.get('student_mobile')
        email_filter = request.POST.get('student_email')

        # Base student query
        student_query = """
            SELECT 
                s.first_name, 
                s.last_name, 
                s.enrollment_id, 
                s.email, 
                s.mobile, 
                s.enrolled_on, 
                COALESCE(rpd.register_payment_status, 'Not Available') AS register_payment_status, 
                COALESCE(rpd.payment_mode, 'Not Available') AS payment_mode, 
                c.name AS counselor_name
            FROM 
                counselor_app_studentenrollment AS s
            JOIN 
                counselor_app_role AS c ON s.counselor_id = c.id
            LEFT JOIN 
                counselor_app_registrationpaymentdetails AS rpd ON s.enrollment_id = rpd.enrollment_id
            WHERE 
                c.manager_id = %s 
                AND c.role_type = 'counselor'
        """
        params = [manager_id]

        # Apply filters
        if course_name_filter:
            student_query += " AND s.course_name = %s"
            params.append(course_name_filter)
        if enrollment_id_filter:
            student_query += " AND s.enrollment_id LIKE %s"
            params.append(f"%{enrollment_id_filter}%")
        if counselor_name_filter:
            student_query += " AND c.name LIKE %s"
            params.append(f"%{counselor_name_filter}%")
        if mobile_filter:
            student_query += " AND s.mobile LIKE %s"
            params.append(f"%{mobile_filter}%")
        if email_filter:
            student_query += " AND s.email LIKE %s"
            params.append(f"%{email_filter}%")

        # Date filter logic
        today = datetime.now().date()
        if date_filter:
            if date_filter == 'today':
                start_date, end_date = today, today + timedelta(days=1)
            elif date_filter == 'yesterday':
                start_date, end_date = today - timedelta(days=1), today
            elif date_filter == 'last_3_days':
                start_date = today - timedelta(days=3)
            elif date_filter == 'last_7_days':
                start_date = today - timedelta(days=7)
            elif date_filter == 'last_month':
                start_date = today - timedelta(days=30)
            elif date_filter == 'custom_date':
                start_date = request.POST.get('start_date')
                end_date = request.POST.get('end_date')
                if start_date and end_date:
                    start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
                    end_date = datetime.strptime(end_date, "%Y-%m-%d").date() + timedelta(days=1)
                else:
                    messages.error(request, "Please provide both start and end dates for custom date filter.")
                    return render(request, 'manager/manager_dashboard_page.html', {'manager_details': manager_details})

            if start_date:
                student_query += " AND s.enrolled_on >= %s"
                params.append(start_date)
            if end_date:
                student_query += " AND s.enrolled_on < %s"
                params.append(end_date)

        # Fetch student data
        with connection.cursor() as cur:
            cur.execute(student_query, params)
            student_data = cur.fetchall()

        student_list = [{
            "s_no": sno + 1,
            "first_name": student[0],
            "last_name": student[1],
            "enrollment_id": student[2],
            "email": student[3],
            "mobile": student[4],
            "enrolled_on": student[5],
            "reg_payment_status": student[6],
            "payment_mode": student[7],
            "counselor_name": student[8]
        } for sno, student in enumerate(student_data)]

        # Paginate the student list
        paginator = Paginator(student_list, 3)  # 3 students per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Get total count of enrolled students
        student_enrolled_count = len(student_list)

        # Fetch course-wise count of students
        with connection.cursor() as cur:
            course_count_query = """
                SELECT c.name AS course_name, COUNT(s.enrollment_id) AS enrollment_count
                FROM counselor_app_studentenrollment AS s
                JOIN counselor_app_role AS c ON s.counselor_id = c.id
                WHERE c.manager_id = %s AND c.role_type = 'counselor'
                GROUP BY c.name
            """
            cur.execute(course_count_query, [manager_id])
            course_data = cur.fetchall()

        course_counts = {course[0]: course[1] for course in course_data}
        course_labels = list(course_counts.keys())
        course_data = list(course_counts.values())

        return render(request, 'manager/manager_dashboard.html', {
            'manager_details': manager_details,
            'student_list': page_obj,
            'student_enrolled_count': student_enrolled_count,
            'course_labels': course_labels,
            'course_data': course_data,
            'course_name_filter': course_name_filter,
            'date_filter': date_filter,
            'enrollment_id_filter': enrollment_id_filter,
            'counselor_name_filter': counselor_name_filter,
            'mobile_filter': mobile_filter,
            'email_filter': email_filter
        })

    except Exception as e:
        print(f"Error occurred in manager_dashboard_page: {e}")
        messages.error(request, f"An error occurred: {e}")
        return redirect('manager_dashboard_page')


def add_role(request):
    try:
        manager_id = request.session.get('manager_id')
        if not manager_id:
            return redirect('manager_login')

        if request.method == 'POST':
            name = request.POST.get('name')
            email = request.POST.get('email')
            employee_id = request.POST.get('employee_id').upper()
            role_type = request.POST.get('role_type')
            added_on = datetime.now()

            if not name or not email or not employee_id or not role_type:
                error = "All fields are required!"
                return render(request, 'manager/add_role.html', {'error': error})

            with connection.cursor() as cur:
                cur.execute("SELECT 1 FROM counselor_app_role WHERE email = %s", [email])
                existing_email = cur.fetchone()
                
                cur.execute("SELECT 1 FROM counselor_app_role WHERE employee_id = %s", [employee_id])
                existing_employee_id = cur.fetchone()

            if existing_email:
                error = "The email is already associated with another role."
                return render(request, 'manager/add_role.html', {'error': error})

            if existing_employee_id:
                error = "The employee ID is already associated with another role."
                return render(request, 'manager/add_role.html', {'error': error})

            with connection.cursor() as cur:
                query = """
                    INSERT INTO counselor_app_role 
                    (manager_id, name, email, employee_id, role_type, added_on)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """
                cur.execute(query, [manager_id, name, email, employee_id, role_type, added_on])

            messages.success(request, "Successfully added role!")
            return redirect('view_team')

        return render(request, 'manager/add_role.html')

    except Exception as e:
        print(f"Error occurred in add_role: {e}")
        messages.error(request, f"An error occurred while adding the role: {e}")
        return render(request, 'manager/add_role.html')

def edit_role_type(request, role_id):
    # Check if the manager is logged in by fetching the manager_id from the session
    manager_id = request.session.get('manager_id')
    if not manager_id:
        return redirect('manager_login')  # Redirect to login page if not logged in

    # Fetch the employee details to edit
    with connection.cursor() as cur:
        # Fetch current employee data based on role_id and manager_id
        employee_query = """
            SELECT id, name, email, role_type, employee_id
            FROM counselor_app_role
            WHERE id = %s AND manager_id = %s
        """
        cur.execute(employee_query, [role_id, manager_id])
        employee_data = cur.fetchone()

    # Check if the employee exists and belongs to this manager
    if not employee_data:
        messages.error(request, "Employee not found or you don't have permission to edit.")
        return redirect('view_team')

    # Process the form submission
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        employee_id = request.POST.get('employee_id')
        new_role_type = request.POST.get('role_type')

        # List of valid role types
        valid_roles = ['counselor', 'developer', 'analyst', 'java_analyst', 'python_analyst', 'trainer']

        # Validate the new role type
        if new_role_type not in valid_roles:
            messages.error(request, "Invalid role type.")
            return redirect('edit_role_type', role_id=role_id)

        # Update the employee data using a raw MySQL query
        try:
            with connection.cursor() as cur:
                update_query = """
                    UPDATE counselor_app_role
                    SET name = %s, email = %s, role_type = %s, employee_id = %s
                    WHERE id = %s AND manager_id = %s
                """
                cur.execute(update_query, [name, email, new_role_type, employee_id, role_id, manager_id])

            messages.success(request, "Employee details updated successfully.")
            return redirect('view_team')

        except Exception as e:
            messages.error(request, f"Error updating employee: {str(e)}")
            return redirect('edit_role_type', role_id=role_id)

    # Render the edit form with current employee data
    return render(request, 'manager/edit_role.html', {
        'employee': {
            'id': employee_data[0],
            'name': employee_data[1],
            'email': employee_data[2],
            'role_type': employee_data[3],
            'employee_id': employee_data[4],
            'role_id': role_id
        }
    })


# def edit_role_type(request, role_id):
#     # Check if the manager is logged in by fetching the manager_id from the session
#     manager_id = request.session.get('manager_id')
#     if not manager_id:
#         return redirect('manager_login')  # Redirect to login page if not logged in

#     # Fetch current employee data based on role_id and manager_id
#     with connection.cursor() as cur:
#         employee_query = """
#             SELECT id, name, email, role_type, employee_id
#             FROM counselor_app_role
#             WHERE id = %s AND manager_id = %s
#         """
#         cur.execute(employee_query, [role_id, manager_id])
#         employee_data = cur.fetchone()
#         print(employee_data)

#         # If no employee data is found, return an error message
#         if not employee_data:
#             return JsonResponse({'status': 'error', 'message': 'Employee not found or unauthorized access.'})

#     # If the request method is POST, update the role_type
#     if request.method == 'POST':
#         new_role_type = request.POST.get('role_type')

#         # Validate the input
#         if not new_role_type:
#             return JsonResponse({'status': 'error', 'message': 'Role type is required.'})

#         # Update the role_type using raw MySQL query
#         try:
#             with connection.cursor() as cur:
#                 update_query = """
#                     UPDATE counselor_app_role
#                     SET role_type = %s
#                     WHERE employee_id = %s AND manager_id = %s
#                 """
#                 cur.execute(update_query, [new_role_type, role_id, manager_id])

#             return JsonResponse({'status': 'success', 'message': 'Role type updated successfully.'})
#         except Exception as e:
#             return JsonResponse({'status': 'error', 'message': str(e)})

#     # Render the edit form with current employee data if the request is GET
#     return render(request, 'manager/edit_role.html', {
#         'employee': {
#             'id': employee_data[0],
#             'name': employee_data[1],
#             'email': employee_data[2],
#             'role_type': employee_data[3],
#             'employee_id':employee_data[4],
#             'role_id': role_id
#         }
#     })

def view_team(request):
    # Check if manager is logged in by fetching the manager_id from session
    manager_id = request.session.get('manager_id')
    if not manager_id:
        return redirect('manager_login')  # Redirect to login page if manager is not logged in

    # Using raw SQL queries to fetch manager and employee details
    with connection.cursor() as cur:
        # Fetching manager details
        manager_query = "SELECT name, email, employee_id FROM counselor_app_manager WHERE id = %s"
        cur.execute(manager_query, [manager_id])
        manager_data = cur.fetchone()

        # Fetching employee details associated with this manager
        employee_query = "SELECT name, email, employee_id, role_type, id FROM counselor_app_role WHERE manager_id = %s"
        cur.execute(employee_query, [manager_id])
        employee_data = cur.fetchall()

    if manager_data:
        # Prepare manager and employee data for rendering
        manager_details = {
            "manager_name": manager_data[0],
            "manager_email": manager_data[1],
            "manager_empid": manager_data[2]
        }

        # Add S.No. (Serial Number) to employee list
        employee_list = [
            {
                "sno": index + 1,  # Serial number starts from 1
                "name": emp[0],
                "email": emp[1],
                "employee_id": emp[2],
                "role_type": emp[3],
                "role_id": emp[4]
            }
            for index, emp in enumerate(employee_data)
        ]

        no_employee_data = not employee_list

        return render(
            request,
            'manager/view_team.html',
            {
                'manager_details': manager_details,
                'employee_list': employee_list,
                'no_employee_data': no_employee_data
            }
        )

# def view_team(request):
#     manager_id = request.session.get('manager_id')
#     if not manager_id:
#         return redirect('manager_login')

#     try:
#         with connection.cursor() as cur:
#             manager_query = "SELECT name, email, employee_id FROM counselor_app_manager WHERE id = %s"
#             cur.execute(manager_query, [manager_id])
#             manager_data = cur.fetchone()

#             employee_query = "SELECT name, email, employee_id, role_type FROM counselor_app_role WHERE manager_id = %s"
#             cur.execute(employee_query, [manager_id])
#             employee_data = cur.fetchall()

#         if manager_data:
#             manager_details = {
#                 "manager_name": manager_data[0],
#                 "manager_email": manager_data[1],
#                 "manager_empid": manager_data[2]
#             }

#             employee_list = [
#                 {
#                     "name": employee[0], 
#                     "email": employee[1], 
#                     "employee_id": employee[2],
#                     "role_type": employee[3]
#                 } 
#                 for employee in employee_data
#             ]

#             no_employee_data = not employee_list

#             return render(request, 'manager/view_team.html', {
#                 'manager_details': manager_details,
#                 'employee_list': employee_list,
#                 'no_employee_data': no_employee_data
#             })
#         else:
#             messages.error(request, "Manager details not found.")
#             return redirect('manager_dashboard_page')

#     except DatabaseError as e:
#         messages.error(request, f"Database error occurred: {str(e)}")
#         return redirect('manager_dashboard_page')

#     except IntegrityError as e:
#         messages.error(request, f"Integrity error occurred: {str(e)}")
#         return redirect('manager_dashboard_page')

#     except Exception as e:
#         messages.error(request, f"An unexpected error occurred: {str(e)}")
#         return redirect('manager_dashboard_page')



def manager_logout(request):
    try:
        manager_id = request.session.get('manager_id')
        if manager_id:
            request.session.pop('manager_id', None)
            messages.success(request, "Successfully logged out!")
        else:
            messages.error(request, "No active session found.")

        return redirect('manager_login')

    except Exception as e:
        messages.error(request, f"An error occurred while logging out: {str(e)}")
        return redirect('manager_dashboard_page')

def add_enroll_students(request):
    emp_id = request.session.get('emp_role_id')
    if not emp_id:
        return render(request, 'employee/employee_login.html')

    if request.method == 'POST':
        try:
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            email = request.POST.get('email')
            mobile = request.POST.get('mobile')
            location = request.POST.get('location')
            mode_of_attending = request.POST.get('mode_of_attending')
            qualification = request.POST.get('qualification')
            branch = request.POST.get('branch')
            course_name = request.POST.get('course_name')
            course_amount = request.POST.get('course_amount')
            discount_amount = request.POST.get('discount_amount')
            total_amount = request.POST.get('total_amount')
            gender = request.POST.get('gender')
            education_status = request.POST.get('education_status')
            passed_year = request.POST.get('passed_year')
            marks = request.POST.get('marks')
            current_year = request.POST.get('current_year')
            enrolled_on = datetime.now()
            enrollment_id = generate_enrollment_id(course_name, enrolled_on)
            registration_fee = request.POST.get('registration_fee')

            with connection.cursor() as cur:
                manager_query = "SELECT manager_id FROM counselor_app_role WHERE id = %s"
                cur.execute(manager_query, [emp_id])
                manager_data = cur.fetchone()

            if manager_data:
                manager_id = manager_data[0]

                with connection.cursor() as cur:
                    check_manager_query = "SELECT id FROM counselor_app_manager WHERE id = %s"
                    cur.execute(check_manager_query, [manager_id])
                    if not cur.fetchone():
                        messages.error(request, "Manager ID does not exist in the manager table.")
                        return render(request, 'student/add_enroll_student.html')

                with connection.cursor() as cur:
                    check_enrollment_query = """
                        SELECT c.name FROM counselor_app_studentenrollment se
                        JOIN counselor_app_role c ON se.counselor_id = c.id
                        WHERE se.email = %s AND se.mobile = %s AND se.course_name = %s
                    """
                    cur.execute(check_enrollment_query, [email, mobile, course_name])
                    existing_enrollment = cur.fetchone()

                    if existing_enrollment:
                        counselor_name = existing_enrollment[0]
                        messages.error(request, f"Student already enrolled by counselor: {counselor_name}.")
                        return render(request, 'student/add_enroll_student.html')

                with connection.cursor() as cur:
                    check_mobile_query = "SELECT * FROM counselor_app_studentenrollment WHERE mobile = %s"
                    cur.execute(check_mobile_query, [mobile])
                    if cur.fetchone():
                        messages.error(request, "Mobile number already exists in the enrollment records.")
                        return render(request, 'student/add_enroll_student.html')

                with connection.cursor() as cur:
                    student_query = """
                        INSERT INTO counselor_app_studentenrollment 
                        (first_name, last_name, email, mobile, location, mode_of_attending, qualification,
                            branch, course_name, course_amount, discount_amount, total_amount, gender,
                            education_status, passed_year, marks, current_year, enrolled_on, enrollment_id,
                            manager_id, counselor_id, registration_fee, status) VALUES 
                        (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s)
                    """
                    cur.execute(student_query, [
                        first_name, last_name, email, mobile, location, mode_of_attending,
                        qualification, branch, course_name, course_amount, discount_amount,
                        total_amount, gender, education_status, passed_year, marks,
                        current_year, enrolled_on, enrollment_id, manager_id, emp_id,
                        registration_fee, 1
                    ])

                messages.success(request, "Student enrolled successfully!")
                return redirect('employee_dashboard_page')

            else:
                messages.error(request, "Manager ID not found.")
                return render(request, 'student/add_enroll_student.html')

        except DatabaseError as db_err:
            messages.error(request, f"Database error: {str(db_err)}")
            return render(request, 'student/add_enroll_student.html')

        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            return render(request, 'student/add_enroll_student.html')

    return render(request, 'student/add_enroll_student.html')

def edit_enrolled_student(request, student_id):
    emp_id = request.session.get('emp_role_id')
    if not emp_id:
        return redirect('employee_login')
    try:
        with connection.cursor() as cur:
            student_query = """
                SELECT id, first_name, last_name, email, mobile, location, mode_of_attending,
                       qualification, branch, course_name, course_amount, discount_amount,
                       total_amount, gender, education_status, passed_year, marks,
                       current_year, enrolled_on, enrollment_id, manager_id, counselor_id
                FROM counselor_app_studentenrollment
                WHERE id = %s
            """
            cur.execute(student_query, [student_id])
            student = cur.fetchone()

        if student:
            student_id, first_name, last_name, email, mobile, location, mode_of_attending, \
            qualification, branch, course_name, course_amount, discount_amount, total_amount, \
            gender, education_status, passed_year, marks, current_year, enrolled_on, \
            enrollment_id, manager_id, counselor_id = student

            if request.method == 'POST':
                updated_first_name = request.POST.get('first_name')
                updated_last_name = request.POST.get('last_name')
                updated_email = request.POST.get('email')
                updated_mobile = request.POST.get('mobile')
                updated_location = request.POST.get('location')
                updated_mode_of_attending = request.POST.get('mode_of_attending')
                updated_qualification = request.POST.get('qualification')
                updated_branch = request.POST.get('branch')
                updated_course_name = request.POST.get('course_name')
                updated_course_amount = request.POST.get('course_amount')
                updated_discount_amount = request.POST.get('discount_amount')
                updated_total_amount = request.POST.get('total_amount')
                updated_gender = request.POST.get('gender')
                updated_education_status = request.POST.get('education_status')
                updated_passed_year = request.POST.get('passed_year')
                updated_marks = request.POST.get('marks')
                updated_current_year = request.POST.get('current_year')

                update_query = """
                    UPDATE counselor_app_studentenrollment
                    SET first_name = %s, last_name = %s, email = %s, mobile = %s,
                        location = %s, mode_of_attending = %s, qualification = %s,
                        branch = %s, course_name = %s, course_amount = %s,
                        discount_amount = %s, total_amount = %s, gender = %s,
                        education_status = %s, passed_year = %s, marks = %s,
                        current_year = %s
                    WHERE id = %s
                """
                with connection.cursor() as cur:
                    cur.execute(update_query, [
                        updated_first_name, updated_last_name, updated_email, updated_mobile,
                        updated_location, updated_mode_of_attending, updated_qualification,
                        updated_branch, updated_course_name, updated_course_amount,
                        updated_discount_amount, updated_total_amount, updated_gender,
                        updated_education_status, updated_passed_year, updated_marks,
                        updated_current_year, student_id
                    ])

                messages.success(request, "Student enrollment updated successfully!")
                return redirect('employee_dashboard_page')

            context = {
                'student': {
                    'id': student_id,
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': email,
                    'mobile': mobile,
                    'location': location,
                    'mode_of_attending': mode_of_attending,
                    'qualification': qualification,
                    'branch': branch,
                    'course_name': course_name,
                    'course_amount': course_amount,
                    'discount_amount': discount_amount,
                    'total_amount': total_amount,
                    'gender': gender,
                    'education_status': education_status,
                    'passed_year': passed_year,
                    'marks': marks,
                    'current_year': current_year,
                },
                'employee_details': request.session.get('employee_details'),
            }
            return render(request, 'student/edit_enroll_student.html', context)

        messages.error(request, "Student not found.")
        return redirect('employee_dashboard_page')

    except DatabaseError as e:
        print(f"Database error: {e}")
        messages.error(request, "An error occurred while retrieving or updating the student data. Please try again.")
        return redirect('employee_dashboard_page')

    except Exception as e:
        print(f"Unexpected error: {e}")
        messages.error(request, "An unexpected error occurred. Please try again.")
        return redirect('employee_dashboard_page')

def delete_enrolled_student(request, student_id):
    emp_id = request.session.get('emp_role_id')
    if not emp_id:
        return redirect('employee_login')
    try:
        with connection.cursor() as cur:
            delete_query = "DELETE FROM counselor_app_studentenrollment WHERE id = %s"
            cur.execute(delete_query, [student_id])

        messages.success(request, "Student data deleted successfully!")
        return redirect('employee_dashboard_page')

    except DatabaseError as e:
        print(f"Database error: {e}")
        messages.error(request, "An error occurred while deleting the student data. Please try again.")
        return redirect('employee_dashboard_page')

    except Exception as e:
        print(f"Unexpected error: {e}")
        messages.error(request, "An unexpected error occurred. Please try again.")
        return redirect('employee_dashboard_page')

def view_enrolled_student(request, student_id):
    emp_id = request.session.get('emp_role_id')
    if not emp_id:
        return redirect('employee_login')

    student_details = {}

    try:
        with connection.cursor() as cur:
            query = """
                SELECT 
                    first_name, last_name, email, course_name, branch,
                    enrollment_id, location, mode_of_attending, qualification,
                    course_amount, discount_amount, total_amount,
                    gender, education_status, passed_year, marks, current_year,
                    enrolled_on 
                FROM counselor_app_studentenrollment 
                WHERE id = %s
            """
            cur.execute(query, [student_id])
            student_data = cur.fetchone()

            if student_data:
                student_details = {
                    'first_name': student_data[0],
                    'last_name': student_data[1],
                    'email': student_data[2],
                    'course_name': student_data[3],
                    'branch': student_data[4],
                    'enrollment_id': student_data[5],
                    'location': student_data[6],
                    'mode_of_attending': student_data[7],
                    'qualification': student_data[8],
                    'course_amount': student_data[9],
                    'discount_amount': student_data[10],
                    'total_amount': student_data[11],
                    'gender': student_data[12],
                    'education_status': student_data[13],
                    'passed_year': student_data[14],
                    'marks': student_data[15],
                    'current_year': student_data[16],
                    'enrolled_on': student_data[17],
                }
                print(student_details)

    except DatabaseError as e:
        print(f"Database error: {e}")
        messages.error(request, "An error occurred while retrieving student data. Please try again.")
        return redirect('employee_dashboard_page')

    except Exception as e:
        print(f"Unexpected error: {e}")
        messages.error(request, "An unexpected error occurred. Please try again.")
        return redirect('employee_dashboard_page')

    return render(request, 'student/view_enrolled_student.html', {'student': student_details})

razorpay_client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_SECRET))

def register_payment_enrolled_student(request, student_id):
    emp_id = request.session.get('emp_role_id')
    if not emp_id:
        return redirect('employee_login')

    student_name = ""
    course_name = ""
    registration_fee = 0
    total_amount = 0
    manager_name = ""
    counselor_name = ""
    enrollment_id = ""

    try:
        with connection.cursor() as cur:
            cur.execute(""" 
                SELECT first_name, last_name, enrollment_id, course_name, total_amount, registration_fee, manager_id 
                FROM counselor_app_studentenrollment 
                WHERE id = %s 
            """, [student_id])
            student_data = cur.fetchone()

            if student_data:
                student_name = f"{student_data[0]} {student_data[1]}"
                enrollment_id = student_data[2]
                course_name = student_data[3]
                total_amount = student_data[4]
                registration_fee = student_data[5]
                manager_id = student_data[6]

                if manager_id:
                    cur.execute("SELECT id, name FROM counselor_app_manager WHERE id = %s", [manager_id])
                    manager_data = cur.fetchone()
                    if manager_data:
                        manager_name = manager_data[1]

                if emp_id:
                    cur.execute("SELECT id, name FROM counselor_app_role WHERE id = %s", [emp_id])
                    counselor_data = cur.fetchone()
                    if counselor_data:
                        counselor_name = counselor_data[1]

        context = {
            'student_id': student_id,
            'student_name': student_name,
            'course_name': course_name,
            'registration_fee': registration_fee,
            'total_amount': total_amount,
            'manager_name': manager_name,
            'counselor_name': counselor_name,
        }

        if request.method == 'POST':
            payment_mode = request.POST.get('payment_mode')
            print(payment_mode)  # Debug print
            with connection.cursor() as cur:
                cur.execute("""
                    INSERT INTO counselor_app_registrationpaymentdetails (manager_id, counselor_id, student_id, 
                    student_name, enrollment_id, course_name, payment_mode, register_payment_status, 
                    paymented_on, payment_by, transaction_id, register_amount, status) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, [
                    manager_id,
                    emp_id,
                    student_id,
                    student_name,
                    enrollment_id,
                    course_name,
                    payment_mode,
                    'success',
                    timezone.now(),
                    counselor_name,
                    "",
                    registration_fee,
                    1
                ])

            messages.success(request, "Payment successfully registered!")
            return redirect('employee_dashboard_page')

    except DatabaseError as e:
        print(f"Database error: {e}")
        messages.error(request, "An error occurred while processing the payment. Please try again.")
        return redirect('employee_dashboard_page')

    except Exception as e:
        print(f"Unexpected error: {e}")
        messages.error(request, "An unexpected error occurred. Please try again.")
        return redirect('employee_dashboard_page')
    return render(request, 'student/register_payment.html', context)



